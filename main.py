import os
import openpyxl
from tabulate import tabulate
import csv
import json


def read_excel(shit):
    workshit = shit
    rows = []
    columns = []
    for row_index in range(workshit.min_row, workshit.max_row + 1):
        for col_index in range(workshit.min_column, workshit.max_column + 1):
            cell = workshit.cell(row=row_index, column=col_index).value
            if cell is None:
                cell = '-'
            rows.append(cell)
        columns.append(rows)
        rows = []
    return columns


def open_excel(file_to_open):
    print('\nOpen_excel working...')
    print('Opening "', file_to_open, '"', sep='')
    workbuk = openpyxl.load_workbook(file_to_open)
    if len(workbuk.sheetnames) == 1:
        print('There are 1 sheet: ', workbuk.sheetnames[0], '"', sep='')
        return read_excel(workbuk[workbuk.sheetnames[0]])
    while True:
        print('There are', len(workbuk.sheetnames), 'sheets:')
        i = 1
        for sheet in workbuk.sheetnames:
            print(i, ': ', sheet, sep='')
            i += 1
        print('Which one do you want to open?')
        print('Enter 1-', i-1, ' to view sheet, "z" to go to main menu or "x" to exit', sep='')
        sel = input()
        try:
            if sel == 'x':
                print('Finishing program...')
                return 'exit'
            elif sel == 'z':
                print('Returning to file selecting...')
                return 'main menu'
            elif 1 <= int(sel) <= len(workbuk.sheetnames):
                print('Selected "', workbuk.sheetnames[int(sel)], '":\n', sep='')
                return read_excel(workbuk[workbuk.sheetnames[int(sel)-1]])
            else:
                print('Value out of range, try again.')
        except ValueError:
            print('Please, enter correct value:')


def menu():
    print('\nCurrent dir is:', os.getcwd())
#    print('Please, enter path to a file (eg: "c:\\games\\jerk-o-tron2021") or "x" to exit:')
    print('Please, enter path to a file (eg: "c:\games\jerk-o-tron2021") or "x" to exit:')
#    path_to_file = 'C:/tpr/githubrep/Javuron'
    while True:
        try:
            path_to_file = input()
            if path_to_file == 'x':
                return 'exit'
            else:
                os.chdir(path_to_file)
                break
        except FileNotFoundError:
            print('Wrong path, try again:')
    print('New dir is:', os.getcwd())
    files = []
    i = 0
    for f in os.listdir(path_to_file):
        if os.path.isfile(path_to_file + '/' + f) is True:
            if (f.count('.xlsx') or f.count('.csv') or f.count('.json') or f.count('.xls')) and not f.count('~'):
                files.append(f)
    print('There are ', len(files), ' files:', sep='')
    for i in range(1, len(files) + 1):
        print(i, ' - ', files[i-1], sep='')
    print('Which one do you want to open?')
    print('Enter 1-', i, ' to select file, or enter "x" to exit, or "z" to change directory', sep='')
    while True:
        x = input()
        try:
            if x == 'x':
                print('Exit')
                return 'exit'
            elif x == 'z':
                print('Main Menu')
                return 'main menu'
            elif 0 < int(x) <= i:
                print('Selected "', files[int(x)-1], '"', sep='')
                return files[int(x)-1]
            else:
                print('Out of range, try again:')
        except ValueError:
            print('Value error, try again:')


def open_csv(file_to_open):
    print('\nOpen_csv working...')
    print('Opening "', file_to_open, '"', sep='')
    with open(file_to_open) as csvfile:
        columns = list(csv.reader(csvfile, delimiter=';'))
        return columns


def open_json(file_to_open):
    print('\nOpen_json working...')
    print('Opening "', file_to_open, '"', sep='')
    tmplist = []
    result = []
    with open(file_to_open, encoding="utf-8") as jsonfile:
        parced = json.load(jsonfile)
        parced_list = parced.get(list(parced.keys())[0])
        result.append(list(parced_list[0].keys()))
        for parced_dicts in parced_list:
            for key in parced_dicts:
                tmplist.append(parced_dicts.get(key))
            result.append(tmplist)
            tmplist = []
        return result


def print_tab(dictionary):
    print(tabulate(dictionary, tablefmt='github'))
    print('\nOpen other file? y/n')
    while True:
        x = input()
        if x == 'y' or x == 'Y':
            return 'main menu'
        elif x == 'n' or x == 'N':
            return 'exit'
        else:
            print('Enter correct value:')
#    print(tabulate(dictionary, headers='firstrow', tablefmt='github'))


def main():
    x = ''
    y = ''
    while x != 'exit':
        x = menu()
        if x.count('.xlsx') or x.count('xls'):
            y = print_tab(open_excel(x))
        elif x.count('.json'):
            y = print_tab(open_json(x))
        elif x.count('.csv'):
            y = print_tab(open_csv(x))
        if y == 'exit':
            return
        elif y == 'main menu':
            pass


main()
